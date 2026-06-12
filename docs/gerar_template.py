"""
Gera o template oficial do PriceLab a partir da base final:
mantém estrutura, cabeçalhos, formatação e premissas; limpa os dados,
deixando 3 linhas de exemplo. Remove abas de resultados (são saída, não entrada).

Uso: python gerar_template.py
"""
import re
import openpyxl

ORIGEM  = r"C:\Users\welli\Downloads\Base Expandir_Final_Dados_Originais.xlsx"
DESTINO = r"C:\Users\welli\OneDrive\Documentos\FAE\Expandir_Framework\backend\app\static\Template_PriceLab.xlsx"

LINHAS_EXEMPLO = 3

wb = openpyxl.load_workbook(ORIGEM)

# 1. Confere que não há fórmulas nas abas de entrada (senão a limpeza quebraria algo)
for nome in ("1_Base_Mensal_v4", "4_Base_Regressao"):
    ws = wb[nome]
    formulas = sum(1 for row in ws.iter_rows() for c in row if c.data_type == "f")
    print(f"{nome}: {formulas} fórmulas encontradas")

# 2. Remove abas de resultado/resumo — template é só entrada de dados
for nome in ("5_Resumo", "3_Resultados_v4"):
    if nome in wb.sheetnames:
        del wb[nome]
        print(f"aba removida: {nome}")

# 3. Limpa dados, mantendo cabeçalhos + linhas de exemplo
def limpar(ws, primeira_linha_dados):
    fim_exemplo = primeira_linha_dados + LINHAS_EXEMPLO - 1
    if ws.max_row > fim_exemplo:
        ws.delete_rows(fim_exemplo + 1, ws.max_row - fim_exemplo)
    print(f"{ws.title}: mantidas linhas 1-{fim_exemplo} (3 exemplos)")

limpar(wb["4_Base_Regressao"], 2)   # linha 1 = header
limpar(wb["1_Base_Mensal_v4"], 3)   # linha 1 = grupos, linha 2 = header

# 3b. Remove fórmulas avulsas que sobraram (ex.: =AVERAGE() de conferência manual)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.data_type == "f":
                print(f"fórmula removida: {ws.title}!{c.coordinate} = {c.value}")
                c.value = None

# 4. Tira o "v4" dos nomes de aba e dos títulos da aba de premissas
_strip_v = lambda s: re.sub(r"\s*[—–-]*\s*v\d+\s*$", "", s, flags=re.I)
for ws in wb.worksheets:
    novo = _strip_v(ws.title).rstrip("_")
    if novo != ws.title:
        print(f"aba renomeada: {ws.title} -> {novo}")
        ws.title = novo

ws_prem = next((ws for ws in wb.worksheets if "premissa" in ws.title.lower()), None)
if ws_prem:
    for row in ws_prem.iter_rows():
        c = row[0]
        if isinstance(c.value, str) and re.search(r"v\d+\s*$", c.value, flags=re.I):
            c.value = _strip_v(c.value)

wb.save(DESTINO)
print(f"\nTemplate salvo em: {DESTINO}")
print("Abas finais:", wb.sheetnames)
