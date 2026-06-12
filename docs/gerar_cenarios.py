"""
Gera duas bases fictícias derivadas da base original para demonstração:
  - Cenário BOM : custos -10%, receitas ajustadas p/ margens ~62% / 55% / +15%
  - Cenário RUIM: custos +18%, receitas ajustadas p/ margens ~5% / -10% / -50%

A estrutura (abas, colunas, formatação) é a da base original; apenas
Receita_Total, Total_Custos, Margem_RS e Margem_PCT são reescalados,
com ruído por linha para manter variação realista.

Uso: python gerar_cenarios.py
"""
import numpy as np
import openpyxl

ORIGEM = r"C:\Users\welli\Downloads\Base Expandir_Final_Dados_Originais.xlsx"
DESTINO_BOM  = r"C:\Users\welli\OneDrive\Documentos\FAE\Expandir_Framework\bases\Base_PriceLab_Cenario_Bom.xlsx"
DESTINO_RUIM = r"C:\Users\welli\OneDrive\Documentos\FAE\Expandir_Framework\bases\Base_PriceLab_Cenario_Ruim.xlsx"

# (fator_custo, {servico: margem_alvo})
CENARIOS = {
    "bom":  (0.90, {"Expansão": 0.62, "Formatação": 0.55, "Validação": 0.15}, DESTINO_BOM,  42),
    "ruim": (1.18, {"Expansão": 0.05, "Formatação": -0.10, "Validação": -0.50}, DESTINO_RUIM, 7),
}


def gerar(nome, fator_custo, alvos, destino, seed):
    rng = np.random.default_rng(seed)
    wb = openpyxl.load_workbook(ORIGEM)

    # Remove abas de resultado (ficariam inconsistentes com os novos valores)
    for aba in ("5_Resumo", "3_Resultados_v4", "1_Base_Mensal_v4"):
        if aba in wb.sheetnames:
            del wb[aba]

    ws = wb["4_Base_Regressao"]
    header = {c.value: c.column for c in ws[1] if c.value}
    col = lambda nome_col: header[nome_col]

    # 1ª passada: razão custo/receita agregada por serviço (na base original)
    soma = {}
    for row in ws.iter_rows(min_row=2):
        sv = row[col("Tipo_Servico") - 1].value
        if not sv:
            continue
        r = float(row[col("Receita_Total") - 1].value or 0)
        c = float(row[col("Total_Custos") - 1].value or 0)
        s = soma.setdefault(sv, [0.0, 0.0])
        s[0] += r
        s[1] += c

    # fator de receita por serviço: margem = 1 - (fc * C) / (k * R)  =>  k = fc*(C/R)/(1-alvo)
    k_receita = {}
    for sv, (r_tot, c_tot) in soma.items():
        alvo = alvos.get(sv, 0.10)
        k_receita[sv] = fator_custo * (c_tot / r_tot) / (1 - alvo)

    # 2ª passada: reescala linha a linha com ruído
    for row in ws.iter_rows(min_row=2):
        sv = row[col("Tipo_Servico") - 1].value
        if not sv:
            continue
        c_rec = row[col("Receita_Total") - 1]
        c_cus = row[col("Total_Custos") - 1]
        c_mrs = row[col("Margem_RS") - 1]

        nova_rec = round(float(c_rec.value) * k_receita[sv] * rng.uniform(0.95, 1.05), 2)
        novo_cus = round(float(c_cus.value) * fator_custo * rng.uniform(0.96, 1.04), 2)
        c_rec.value = nova_rec
        c_cus.value = novo_cus
        c_mrs.value = round(nova_rec - novo_cus, 2)
        if "Margem_PCT" in header:
            row[col("Margem_PCT") - 1].value = round((nova_rec - novo_cus) / nova_rec, 4)

    wb.save(destino)
    print(f"cenário {nome}: salvo em {destino}")


for nome, (fc, alvos, destino, seed) in CENARIOS.items():
    gerar(nome, fc, alvos, destino, seed)
